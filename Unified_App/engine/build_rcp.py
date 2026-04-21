"""
Excel workbook → RCP file builder.
Adapted from: Empirical RCP/build_rcp.py
(Original file not modified.)
"""
from __future__ import annotations

import re
from collections import Counter
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Callable, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook


DEFAULT_DSLIB_SHEET = "DSLIB"
DEFAULT_BASELINE_SHEET = "Baseline Schedule"


def decimal_to_str(value: Decimal) -> str:
    text = format(value, "f").rstrip("0").rstrip(".")
    return text or "0"


def format_num(value: object) -> str:
    if value is None:
        return "0"
    if isinstance(value, Decimal):
        return decimal_to_str(value)
    try:
        num = Decimal(str(value))
    except InvalidOperation:
        return str(value)
    return decimal_to_str(num)


def parse_decimal(text: str, default: Decimal = Decimal("1")) -> Decimal:
    match = re.search(r"[-+]?[0-9]*[.,]?[0-9]+", text)
    if not match:
        return default
    return Decimal(match.group(0).replace(",", "."))


def load_header_map(ws, header_row: int) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        if isinstance(value, str) and value.strip():
            header_map[value.strip()] = col
    return header_map


def find_header_row(ws, required: Iterable[str], search_rows: int = 10) -> int:
    required_set = {name.strip() for name in required}
    for row in range(1, search_rows + 1):
        row_names = {
            ws.cell(row=row, column=col).value
            for col in range(1, ws.max_column + 1)
        }
        row_names_clean = {v.strip() for v in row_names if isinstance(v, str)}
        if required_set.issubset(row_names_clean):
            return row
    raise ValueError(f"Could not find header row containing: {sorted(required_set)}")


def detect_activity_fill(ws, id_col: int, start_row: int) -> Tuple[str, str, str]:
    counter: Counter = Counter()
    for row in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=row, column=id_col)
        fill = cell.fill
        key = (fill.patternType, fill.fgColor.type, fill.fgColor.value)
        counter[key] += 1
    if not counter:
        raise ValueError("No fills found when detecting activity rows.")
    return counter.most_common(1)[0][0]


def parse_resources(text: Optional[str]) -> List[Tuple[str, Decimal]]:
    if not text or not isinstance(text, str):
        return []
    parts = [part.strip() for part in text.split(";") if part.strip()]
    results: List[Tuple[str, Decimal]] = []
    for part in parts:
        match = re.match(r"^(.*?)\s*\[(.*)\]$", part)
        if match:
            name = match.group(1).strip()
            quantity = parse_decimal(match.group(2), default=Decimal("1"))
        else:
            name = part.strip()
            quantity = Decimal("1")
        if name:
            results.append((name, quantity))
    return results


def parse_successors(value: object) -> List[int]:
    if value is None:
        return []
    return [int(m) for m in re.findall(r"\d+", str(value))]


def build_rcp(
    project_path: Path,
    output_path: Path,
    baseline_sheet: str = DEFAULT_BASELINE_SHEET,
) -> Tuple[int, int]:
    """
    Build an RCP file from an Excel project workbook.
    Returns (activity_count, resource_count).
    """
    project_wb = load_workbook(project_path, data_only=True)
    if baseline_sheet not in project_wb.sheetnames:
        available = ", ".join(project_wb.sheetnames)
        raise ValueError(
            f"Sheet '{baseline_sheet}' not found in workbook. Available: {available}"
        )
    baseline_ws = project_wb[baseline_sheet]

    baseline_header_row = find_header_row(
        baseline_ws,
        ["ID", "Successors", "Resource Demand", "Baseline duration (in calendar days)"],
    )
    baseline_header_map = load_header_map(baseline_ws, baseline_header_row)

    id_col = baseline_header_map["ID"]
    resource_col = baseline_header_map["Resource Demand"]
    successors_col = baseline_header_map["Successors"]
    duration_col = baseline_header_map["Baseline duration (in calendar days)"]

    activity_fill = detect_activity_fill(baseline_ws, id_col, baseline_header_row + 1)

    activities = []
    resource_order: List[str] = []
    resource_index: Dict[str, int] = {}

    for row in range(baseline_header_row + 1, baseline_ws.max_row + 1):
        id_cell = baseline_ws.cell(row=row, column=id_col)
        fill = id_cell.fill
        fill_key = (fill.patternType, fill.fgColor.type, fill.fgColor.value)
        if fill_key != activity_fill:
            continue
        duration_value = baseline_ws.cell(row=row, column=duration_col).value
        resources_text = baseline_ws.cell(row=row, column=resource_col).value
        successors_value = baseline_ws.cell(row=row, column=successors_col).value
        resource_demands: Dict[str, Decimal] = {}
        for name, quantity in parse_resources(resources_text):
            if name not in resource_index:
                resource_index[name] = len(resource_order)
                resource_order.append(name)
            resource_demands[name] = resource_demands.get(name, Decimal("0")) + quantity
        activities.append({
            "original_id": id_cell.value,
            "duration": duration_value,
            "resources": resource_demands,
            "successors": parse_successors(successors_value),
        })

    if not activities:
        raise ValueError("No activity rows found. Check that the baseline schedule sheet has the correct format.")

    if not resource_order:
        default_resources = ["Resource_1", "Resource_2", "Resource_3", "Resource_4"]
        for rname in default_resources:
            resource_index[rname] = len(resource_order)
            resource_order.append(rname)
        for activity in activities:
            for rname in default_resources:
                activity["resources"][rname] = Decimal("10")

    id_map = {a["original_id"]: idx + 1 for idx, a in enumerate(activities)}
    resource_count = len(resource_order)
    demands_matrix: List[List[Decimal]] = []
    successors_matrix: List[List[int]] = []

    for activity in activities:
        row_demands = [Decimal("0")] * resource_count
        for name, qty in activity["resources"].items():
            row_demands[resource_index[name]] = qty
        demands_matrix.append(row_demands)
        mapped_succs = [id_map[s] for s in activity["successors"] if s in id_map]
        successors_matrix.append(mapped_succs)

    capacities = [max(col) for col in zip(*demands_matrix)]

    lines: List[str] = [
        "",
        f"{len(activities)} {resource_count}",
        "  " + "  ".join(format_num(v) for v in capacities),
        "",
    ]
    for activity, demands, succs in zip(activities, demands_matrix, successors_matrix):
        parts = [format_num(activity["duration"])]
        parts.extend(format_num(v) for v in demands)
        parts.append(str(len(succs)))
        parts.extend(str(s) for s in succs)
        lines.append("  " + "  ".join(parts))
    terminator = ["0"] + ["0"] * resource_count + ["0"]
    lines.append("  " + "  ".join(terminator))
    lines.append("")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(lines), encoding="ascii")
    return len(activities), resource_count


def run_build_rcp(
    project_path: Path,
    output_dir: Path,
    output_filename: Optional[str] = None,
    baseline_sheet: str = DEFAULT_BASELINE_SHEET,
    progress_cb: Optional[Callable] = None,
    cancel_check: Optional[Callable] = None,
) -> Path:
    """
    Build an RCP file from a project workbook and save to output_dir.
    Returns path to generated RCP file.
    """
    project_path = Path(project_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    if output_filename is None:
        output_filename = project_path.stem + ".rcp"
    if not output_filename.endswith(".rcp"):
        output_filename += ".rcp"

    output_path = output_dir / output_filename

    if progress_cb:
        progress_cb(0, 2, f"Reading workbook: {project_path.name}...")

    if cancel_check and cancel_check():
        raise InterruptedError("RCP build cancelled by user.")

    activity_count, resource_count = build_rcp(project_path, output_path, baseline_sheet)

    if progress_cb:
        progress_cb(2, 2, f"Done. {activity_count} activities, {resource_count} resources → {output_path.name}")

    return output_path
